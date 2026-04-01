from __future__ import annotations

import csv
import json
import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
SONNIES_PATH = ROOT / "app" / "data" / "sonnies.json"
IMAGE_MAP_PATH = ROOT / "app" / "data" / "sonny_image_map.json"
MANUAL_OVERRIDES_PATH = ROOT / "app" / "data" / "manual_overrides.js"
MANIFEST_PATH = ROOT / "sonny_png_library" / "manifest.csv"
OUT_XLSX_PATH = ROOT / "master_sonny_catalog.xlsx"
OUT_CSV_PATH = ROOT / "master_sonny_catalog.csv"

SECRET_IDS = {
    "page-10-13-chipmunk",
}

ACCOUNTED_SERIES = {
    "Animal Series Ver. 1": {
        "status": "ACCOUNTED",
        "expected_count": 14,
        "note": "Refined Animal Series 1 audited and accounted for.",
    },
}

SERIES_ALIASES = {
    "animal series ver. 1": "animal series 1",
    "animal series ver. 2": "animal series 2",
    "animal series ver. 3": "animal series 3",
    "animal series ver. 4": "animal series 4",
    "marine series": "marine series",
    "flower series": "flower series",
    "fruit series": "fruit series",
    "vegetable series": "vegetable series",
    "sweets series": "sweets series",
}

HEADER_FILL = PatternFill("solid", fgColor="EFD7B4")
ACCOUNTED_FILL = PatternFill("solid", fgColor="D8E3CF")
REVIEW_FILL = PatternFill("solid", fgColor="F4C8CF")

ASSIGN_RE = re.compile(r'assign\("([^"]+)",\s*\{(.*?)\}\);', re.S)
FIELD_RE = re.compile(r'(\w+):\s*(true|false|"[^"]*")\s*,?')


def normalize(text: str) -> str:
    return " ".join((text or "").lower().replace("’", "'").split())


def normalize_series(text: str) -> str:
    base = normalize(text)
    return SERIES_ALIASES.get(base, base)


def autosize(ws) -> None:
    widths: dict[int, int] = defaultdict(int)
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column] = max(widths[cell.column], len(str(cell.value)))
    for column, width in widths.items():
        ws.column_dimensions[get_column_letter(column)].width = min(width + 2, 48)


def load_manifest() -> dict[str, dict]:
    rows = list(csv.DictReader(MANIFEST_PATH.open()))
    manifest_by_path: dict[str, dict] = {}
    for row in rows:
        key = row["output_path"].replace("sonny_png_library_web/", "sonny_png_library/")
        manifest_by_path[key] = row
    return manifest_by_path


def load_existing_master_rows() -> dict[str, dict]:
    if not OUT_CSV_PATH.exists():
        return {}
    rows = list(csv.DictReader(OUT_CSV_PATH.open()))
    return {row["id"]: row for row in rows}


def load_manual_overrides() -> dict[str, dict]:
    text = MANUAL_OVERRIDES_PATH.read_text()
    overrides: dict[str, dict] = {}
    for match in ASSIGN_RE.finditer(text):
        item_id, body = match.groups()
        patch: dict[str, object] = {}
        for key, raw_value in FIELD_RE.findall(body):
            if raw_value == "true":
                value: object = True
            elif raw_value == "false":
                value = False
            else:
                value = raw_value[1:-1]
            patch[key] = value
        if patch:
            overrides[item_id] = {**overrides.get(item_id, {}), **patch}
    return overrides


def identity_status_for(
    *,
    series: str,
    name: str,
    mapped_path: str,
    mapped_series: str,
    mapped_name: str,
    existing_row: dict,
) -> str:
    if not mapped_path:
        return "UNMAPPED"
    if mapped_series and mapped_name:
        if normalize_series(mapped_series) == normalize_series(series) and normalize(mapped_name) == normalize(name):
            return "MATCHED"
        return "REVIEW"
    if existing_row.get("identity_status"):
        return existing_row["identity_status"]
    return "MATCHED"


def build_catalog_rows() -> list[dict]:
    sonnies = json.loads(SONNIES_PATH.read_text())
    image_map = json.loads(IMAGE_MAP_PATH.read_text())
    manifest_by_path = load_manifest()
    existing_rows = load_existing_master_rows()
    manual_overrides = load_manual_overrides()

    rows: list[dict] = []
    for entry in sonnies:
        entry_id = entry["id"]
        existing_row = existing_rows.get(entry_id, {})
        override = manual_overrides.get(entry_id, {})
        mapped = image_map.get(entry_id, {})

        series = str(override.get("series") or entry["series"])
        name = str(override.get("name") or entry["name"])
        mapped_path = (
            str(override.get("artPath") or mapped.get("path") or existing_row.get("mapped_asset_path") or "")
            .replace("./", "")
        )
        mapped_series = str(mapped.get("catalogSeries") or existing_row.get("mapped_catalog_series") or "")
        mapped_name = str(mapped.get("catalogName") or existing_row.get("mapped_catalog_name") or "")
        mapped_source = str(override.get("artSource") or mapped.get("source") or existing_row.get("mapped_source") or "")
        identity_status = identity_status_for(
            series=series,
            name=name,
            mapped_path=mapped_path,
            mapped_series=mapped_series,
            mapped_name=mapped_name,
            existing_row=existing_row,
        )
        manifest_row = manifest_by_path.get(mapped_path) if mapped_path else None

        rows.append(
            {
                "id": entry_id,
                "page": entry["page"],
                "series": series,
                "name": name,
                "is_secret": "YES"
                if override.get("isSecret") or entry_id in SECRET_IDS or existing_row.get("is_secret") == "YES"
                else "",
                "db_image_page": entry["image"]["page"],
                "db_x": entry["image"]["x"],
                "db_y": entry["image"]["y"],
                "mapped_asset_path": mapped_path or "",
                "mapped_catalog_series": mapped_series or "",
                "mapped_catalog_name": mapped_name or "",
                "mapped_source": mapped_source or "",
                "identity_status": identity_status,
                "manifest_image_url": manifest_row["image_url"] if manifest_row else existing_row.get("manifest_image_url", ""),
                "manifest_page_url": manifest_row["page_url"] if manifest_row else existing_row.get("manifest_page_url", ""),
            }
        )
    return rows


def build_series_rows(catalog_rows: list[dict]) -> list[dict]:
    by_series: dict[str, list[dict]] = defaultdict(list)
    series_order: list[str] = []
    for row in catalog_rows:
        series = row["series"]
        if series not in by_series:
            series_order.append(series)
        by_series[series].append(row)

    series_rows: list[dict] = []
    for series in series_order:
        rows = by_series[series]
        db_count = len(rows)
        matched_count = sum(1 for row in rows if row["identity_status"] == "MATCHED")
        review_count = sum(1 for row in rows if row["identity_status"] == "REVIEW")
        unmapped_count = sum(1 for row in rows if row["identity_status"] == "UNMAPPED")
        accounted = ACCOUNTED_SERIES.get(series)
        expected_count = accounted["expected_count"] if accounted else db_count
        audit_status = accounted["status"] if accounted else (
            "READY" if review_count == 0 and unmapped_count == 0 else "NEEDS_REVIEW"
        )
        note = accounted["note"] if accounted else ""
        series_rows.append(
            {
                "series": series,
                "db_count": db_count,
                "expected_count": expected_count,
                "matched_count": matched_count,
                "review_count": review_count,
                "unmapped_count": unmapped_count,
                "audit_status": audit_status,
                "note": note,
            }
        )
    return series_rows


def write_catalog_sheet(ws, catalog_rows: list[dict]) -> None:
    headers = list(catalog_rows[0].keys())
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL

    for row in catalog_rows:
        ws.append([row[header] for header in headers])
        status_cell = ws.cell(ws.max_row, headers.index("identity_status") + 1)
        if row["identity_status"] == "MATCHED":
            status_cell.fill = ACCOUNTED_FILL
        elif row["identity_status"] != "UNMAPPED":
            status_cell.fill = REVIEW_FILL

    ws.freeze_panes = "A2"
    autosize(ws)


def write_series_sheet(ws, series_rows: list[dict]) -> None:
    headers = list(series_rows[0].keys())
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL

    for row in series_rows:
        ws.append([row[header] for header in headers])
        status_cell = ws.cell(ws.max_row, headers.index("audit_status") + 1)
        if row["audit_status"] in {"ACCOUNTED", "READY"}:
            status_cell.fill = ACCOUNTED_FILL
        else:
            status_cell.fill = REVIEW_FILL

    ws.freeze_panes = "A2"
    autosize(ws)


def write_notes_sheet(ws) -> None:
    rows = [
        ("Purpose", "Master Sonny catalog generated from the tracker database so each row is a real Sonny entry."),
        ("Database source", str(SONNIES_PATH.relative_to(ROOT))),
        ("Manual overrides source", str(MANUAL_OVERRIDES_PATH.relative_to(ROOT))),
        ("Image map source", str(IMAGE_MAP_PATH.relative_to(ROOT))),
        ("Manifest source", str(MANIFEST_PATH.relative_to(ROOT))),
        ("Catalog order", "Rows follow the exact collection and item order in app/data/sonnies.json."),
        ("Overrides", "Series, names, secrets, and manual art paths from manual_overrides.js are applied before export."),
        ("Accounted now", "Animal Series Ver. 1 (refined) is marked ACCOUNTED with 14/14 entries."),
        ("Review meaning", "REVIEW means the mapped asset identity does not exactly match the database series/name."),
        ("Unmapped meaning", "UNMAPPED means there is no linked asset path in the current image map."),
    ]
    for row in rows:
        ws.append(row)
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 110


def main() -> None:
    catalog_rows = build_catalog_rows()
    series_rows = build_series_rows(catalog_rows)

    wb = Workbook()
    catalog_ws = wb.active
    catalog_ws.title = "catalog"
    write_catalog_sheet(catalog_ws, catalog_rows)

    series_ws = wb.create_sheet("series_audit")
    write_series_sheet(series_ws, series_rows)

    notes_ws = wb.create_sheet("notes")
    write_notes_sheet(notes_ws)

    wb.save(OUT_XLSX_PATH)

    with OUT_CSV_PATH.open("w", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(catalog_rows[0].keys()))
        writer.writeheader()
        writer.writerows(catalog_rows)

    accounted_counter = Counter(row["audit_status"] for row in series_rows)
    print(f"saved {OUT_XLSX_PATH}")
    print(f"saved {OUT_CSV_PATH}")
    print(
        f"series statuses: ACCOUNTED={accounted_counter['ACCOUNTED']} "
        f"READY={accounted_counter['READY']} NEEDS_REVIEW={accounted_counter['NEEDS_REVIEW']}"
    )


if __name__ == "__main__":
    main()
